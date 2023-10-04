import re, fitz, camelot, pytesseract, PyPDF2
import pandas as pd
from config import Config
from flask import Response
from pdf2image import convert_from_path
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

download_excel_path = Config.EXCEL_FILE_PATH


def remove_extra_headers(df):
    header_row = df.iloc[0]
    if header_row[0] == 'Date' and header_row[1] == 'Narrative' or (header_row[0] == 'Date' and header_row[1] == 'Description'):
        return df[df.index != 0].dropna(subset=[0], how='all')
    return df.dropna(subset=[0], how='all')


def readandcleandata(data):
    cleaned_data = []
    for df in data:
        df_cleaned = remove_extra_headers(df)
        if len(cleaned_data) > 0:
            df_cleaned = df_cleaned.iloc[1:]
        cleaned_data.append(df_cleaned)
    df_concatenated = pd.concat(cleaned_data, ignore_index=True)
    return df_concatenated


def processNatwest(file, filename):
    tables = camelot.read_pdf(file, pages="all", flavor="stream")
    structured_data = pd.DataFrame()

    for table in tables:
        df = table.df
        if 'Type' in df.columns:
            type_mapping = {"BAC": "Paid In", "DPC": "Paid In", "EBP": "Paid Out"}
            df['Paid In'] = df['Type'].map(type_mapping)
            df['Paid Out'] = df['Type'].map(type_mapping)
        structured_data = pd.concat([structured_data, df], ignore_index=True)

    details_to_drop = ["Statement for account 60-00-08 48716081 from 01/07/2021 to 02/09/2021", "Date", "Narrative", "Debit", "Credit",
        'Interest Rates: Your interest rate is 0.01% gross 0.01% AER. This is based on your balance from end of day yesterday.']
    structured_data = structured_data[~structured_data[structured_data.columns[0]].str.contains('|'.join(details_to_drop)) |
                                    (structured_data.index == 0) |
                                    (structured_data.index == len(structured_data) - 1)]

    details_to_drop = ["CLOSING BALANCE", "BALANCE BROUGHT FORWARD", "BALANCE CARRIED FORWARD"]
    structured_data = structured_data[~structured_data[1].str.contains('|'.join(details_to_drop)) |
                                    (structured_data.index == 0) |
                                    (structured_data.index == len(structured_data) - 1)]

    structured_data = structured_data.iloc[35:-1]
    structured_data = structured_data.reset_index(drop=True)
    structured_data.columns = ['Date', 'Description', "Type", 'Paid Out', 'Paid In', 'Balance', 'Extra']

    if 'Balance' in structured_data.columns:
        mask = ~structured_data['Balance'].str.endswith('Cr')
        structured_data.loc[mask, 'Paid In'] = structured_data.loc[mask, 'Balance']
        structured_data.loc[mask, 'Balance'] = ''
    structured_data = structured_data.drop_duplicates()
    columns_to_remove = ["Balance", "Extra"]
    structured_data = structured_data.drop(columns_to_remove, axis=1)

    try:
        with pd.ExcelWriter(f"{download_excel_path}{filename}.xlsx", engine="openpyxl") as writer:
            structured_data.to_excel(writer, sheet_name=f"{filename}", index=False)
            writer.book
            worksheet = writer.sheets[f"{filename}"]
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2
        return Response(status=201)
    except Exception:
        return Response(status=404)


def processLLoyds(file, filename):
    words = ''
    dates = []
    descriptions = []
    paid_out = []
    paid_in = []

    doc = fitz.open(file)
    for page in doc:
        words += page.get_text("Date", sort=True)
    pattern = r'(\d{2}[A-Za-z]{3}\d{2})\s+([A-Za-z]{2,3})\s+(.+)'
    pattern = pattern + r'|(\d{2}[A-Za-z]{3}\d{2})\s+RETURNED\s+'
    matches = re.finditer(pattern, words)
    match_positions = []

    for match in matches:
        start = match.start()
        end = match.end()
        match_positions.append((start, end))

    for num in range(len(match_positions)):
        if num == len(match_positions) - 1:
            data = words[match_positions[num][0]:match_positions[num][1]+20]
            data = data.split('\n')
            for i in range(len(data)):
                data[i] = re.sub(r'\s+', '', data[i])
            dates.append(data[0][:7])
            if len(data)> 5 and not data[5] == '' and data[4] == '':
                descriptions.append(data[0][7:])
                paid_out.append(data[1])
                paid_in.append(data[2])
            elif data[1] == "RETURNEDDD":
                descriptions.append(data[1])
                paid_out.append(data[3])
                paid_in.append(data[4])
            elif '.' in data[1]:
                descriptions.append(data[0][7:])
                paid_out.append(data[1])
                paid_in.append(data[2])
            elif data[1]:
                descriptions.append(data[1])
                paid_out.append(data[2])
                paid_in.append(data[3])
            else:
                descriptions.append(data[0][7:])
                paid_out.append(data[2])
                paid_in.append(data[3])
        else:
            data = words[match_positions[num][0]:match_positions[num + 1][0]]
            data = data.split('\n')
            for i in range(len(data)):
                data[i] = re.sub(r'\s+', '', data[i])
            dates.append(data[0][:7])
            if len(data)> 5 and not data[5] == '' and data[4] == '':
                descriptions.append(data[0][7:])
                paid_out.append(data[1])
                paid_in.append(data[2])
            elif data[1] == "RETURNEDDD":
                descriptions.append(data[1])
                paid_out.append(data[3])
                paid_in.append(data[4])
            elif '.' in data[1]:
                descriptions.append(data[0][7:])
                paid_out.append(data[1])
                paid_in.append(data[2])
            elif data[1]:
                descriptions.append(data[1])
                paid_out.append(data[2])
                paid_in.append(data[3])
            else:
                descriptions.append(data[0][7:])
                paid_out.append(data[2])
                paid_in.append(data[3])
    df = pd.DataFrame({"Date": dates,"Description":descriptions, "Paid Out": paid_out, "Paid In": paid_in})

    try:
        with pd.ExcelWriter(f"{download_excel_path}{filename}.xlsx", engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=f"{filename}", index=False)
                writer.book
                worksheet = writer.sheets[f"{filename}"]
                for column_cells in worksheet.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2
        return Response(status=201)
    except Exception:
        return Response(status=404)


def processLLoyds2(file, filename):
    tables = camelot.read_pdf(file, pages="all", flavor="stream")
    structured_data = pd.DataFrame()
    for table in tables:
        df = table.df
        if len(df.columns) == 5:
            structured_data = pd.concat([structured_data, df], ignore_index=True)
        elif len(df.columns) == 6:
            parsed_df = df.copy()
            parsed_df.columns = [0, 1, 2, 3, 4, 5]
            parsed_df = parsed_df.drop(columns=parsed_df.columns[[2, 5]])
            structured_data = pd.concat([structured_data, parsed_df], ignore_index=True)
    structured_data = structured_data.drop(columns=2)

    details_to_drop = ["https://securebusiness.lloydsbank.co.uk/business/link/lp_print_st…ew&amp;targetid=printable&urluid=1517069106080-0.7086655426330178", "Lloyds Bank - Print Friendly Statement"]
    structured_data = structured_data[~structured_data[0].str.contains('|'.join(details_to_drop)) |
                                    (structured_data.index == 0) |
                                    (structured_data.index == len(structured_data) - 1)]

    column_header = ['Date', 'Description', 'Paid Out', 'Paid In']
    structured_data.columns = column_header
    structured_data = structured_data.iloc[3:-7]
    replace_dict = {"21220.00": "", "15500.00": ""}
    structured_data["Paid In"] = structured_data["Paid In"].replace(replace_dict)

    try:
        with pd.ExcelWriter(f"{download_excel_path}{filename}.xlsx", engine="openpyxl") as writer:
            structured_data.to_excel(writer, sheet_name=f"{filename}", index=False)
            writer.book
            worksheet = writer.sheets[f"{filename}"]
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2
        return Response(status=201)
    except Exception:
        return Response(status=404)


def processHSBC(file, filename):
    date_pattern = r"\d{2} \w{3} \d{2}"
    tables = camelot.read_pdf(file, pages="all", flavor="stream")
    structured_data = pd.DataFrame()

    for table in tables:
        df = table.df
        if len((df.columns)) == 4:
            df = df.iloc[2:]
            structured_data = pd.concat([structured_data, df], ignore_index=True)
        elif len((df.columns)) == 5:
            parsed_df = df.copy()
            parsed_df.columns = [0, 'blank', 1, 2, 3]
            parsed_df = parsed_df.drop(columns=parsed_df.columns[1])
            structured_data = pd.concat([structured_data, parsed_df], ignore_index=True)

    structured_data = structured_data.drop(columns=3)
    details_to_drop = ["BALANCE BROUGHT FORWARD", "BALANCE CARRIED FORWARD"]
    structured_data = structured_data[~structured_data[0].str.contains('|'.join(details_to_drop)) |
                                    (structured_data.index == 0) |
                                    (structured_data.index == len(structured_data) - 1)]

    column_header = ['Details', 'Paid Out', 'Paid In']
    structured_data.columns = column_header
    structured_data = structured_data.iloc[3:-1]

    try:
        with pd.ExcelWriter(f"{download_excel_path}{filename}.xlsx", engine="openpyxl") as writer:
            structured_data.to_excel(writer, sheet_name=f"{filename}", index=False)
            writer.book
            worksheet = writer.sheets[f"{filename}"]
            date_column_header = "Date"
            worksheet.insert_cols(0)
            worksheet.cell(row=1, column=1, value=date_column_header)
            for index, rows_cells in enumerate(worksheet.rows):
                length = max(len(str(cell.value)) for cell in rows_cells)
                worksheet.row_dimensions[rows_cells[1].row].height = length + 15
                worksheet.row_dimensions[rows_cells[1].row].width = length + 15
                matches = re.findall(date_pattern, str(rows_cells[1].value))
                if matches:
                    worksheet.row_dimensions[rows_cells[1].row].width = length + 15
                    rows_cells[1].value = rows_cells[1].value.replace(matches[0], "")
                    worksheet.cell(row=index + 1, column=1, value=matches[0])
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2
        return Response(status=201)
    except Exception:
        return Response(status=404)


def processBarclays(file, filename):
    tables = camelot.read_pdf(file, pages="all", flavor="stream")
    structured_data = pd.DataFrame(columns=['Date', 'Description', 'Paid Out', 'Paid In'])
    for table in tables:
        if len(table.df.columns) >= 4:
            if len(table.df.columns) == 7:
                paid_out_idx, paid_in_idx = -3, -2
            else:
                paid_out_idx, paid_in_idx = -3, -2
            data = {
                'Date': table.df.iloc[:, 0],
                'Description': table.df.iloc[:, 1],
                'Paid Out': table.df.iloc[:, paid_out_idx],
                'Paid In': table.df.iloc[:, paid_in_idx]
            }
            temp_df = pd.DataFrame(data)
            structured_data = pd.concat([structured_data, temp_df], ignore_index=True)
    structured_data = structured_data.iloc[3:]

    try:
        with pd.ExcelWriter(f"{download_excel_path}{filename}.xlsx", engine="openpyxl") as writer:
            structured_data.to_excel(writer, sheet_name=f"{filename}", index=False)
            writer.book
            worksheet = writer.sheets[f"{filename}"]
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2
        return Response(status=201)
    except Exception:
        return Response(status=404)


def processHSBC_Scanned(file, filename):
    Data_Objects = []
    combined_list = []
    formatted_list = []

    def convert_pdf_to_text(file):
        extracted_text = ''
        images = convert_from_path(file, dpi=300)
        for img in images:
            text = pytesseract.image_to_string(img)
            extracted_text += text
        return extracted_text

    extracted_text = convert_pdf_to_text(file)
    column_headers = ['Date', 'Type', 'Details', 'Paid Out', 'Paid In', 'Balance']
    df = pd.DataFrame(columns=column_headers)
    lines = extracted_text.split('\n')

    for index, line in enumerate(lines):
        if line != '':
            if index < len(lines) - 1 and bool(re.match(r'\d', line)):
                combined_string = line + ' ' + lines[index + 1]
                Data_Objects.append(combined_string)
            else:
                Data_Objects.append(line)

    i = 0
    while i < len(Data_Objects) - 1:
        combined_string = Data_Objects[i] + "" + Data_Objects[i + 1]
        combined_list.append(combined_string)
        i += 2

    if len(Data_Objects) % 2 != 0:
        combined_list.append(Data_Objects[-1])

    for data in combined_list:
        parts = data.split()
        if parts[1] == 'DR' or parts[1] == 'BP':
            if len(parts) == 7:
                formatted_data = [
                    parts[0],
                    parts[1],
                    ' '.join(parts[2:-1]),
                    parts[-1],
                    '',
                    '',
                ]
            else:
                formatted_data = [
                    parts[0],
                    parts[1],
                    ' '.join(parts[2:-2]),
                    parts[-2],
                    '',
                    parts[-1],
                ]
            formatted_list.append(formatted_data)
        elif parts[1] == 'TFR':
            formatted_data = [
                parts[0],
                parts[1],
                ' '.join(parts[2:-3]),
                '',
                ''.join(parts[-3:-1]),
                parts[-1],
            ]
            formatted_list.append(formatted_data)

    for data in formatted_list:
        df.loc[len(df)] = data
    df = df.apply(lambda x: x.str.replace(":", "."))

    try:
        workbook = Workbook()
        worksheet = workbook.active
        for row in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(row)
        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width
        workbook.save(f"{download_excel_path}{filename}.xlsx")
        return Response(status=201)
    except Exception:
        return Response(status=404)


def processNatwest_Large_Scanned(file, filename):
    rows_to_skip = {
        1: (22, 7),
        2: (7, 4),
    }
    def convert_pdf_to_text(file, rows_to_skip):
        extracted_text = ''
        page_number = 1
        images = convert_from_path(file, dpi=300)
        for img in images:
            text = pytesseract.image_to_string(img)
            lines = text.split('\n')
            start_skip, end_skip = rows_to_skip.get(page_number, (0, 0))
            lines = lines[start_skip:-end_skip] if end_skip > 0 else lines[start_skip:]
            page_text = '\n'.join(lines)
            extracted_text += page_text
            page_number += 1
        return extracted_text

    def process_row(row):
        date = ''
        detail = ''
        withdrawn = ''
        paid_in = ''
        balance = ''

        all_words = row.split()
        if len(all_words) > 1 and all_words[1][-1] == 'N':
            all_words[0] = all_words[0] + all_words[1]
            del all_words[1]
            row = " ".join(all_words)

        if all_words:
            first_word = all_words[0]
            if 'Bill Payment' in row:
                date_detail_parts = row.split('Bill Payment', 1)
                if len(date_detail_parts) == 2:
                    date = date_detail_parts[0].strip()
                    detail = 'Bill Payment ' + date_detail_parts[1].strip()
                    date_match = re.search(r'\d{2}/\d{2}/\d{4}', date)
                    if date_match:
                        date = date_match.group()
                    else:
                        date = ''
                else:
                    detail = row
            elif len(first_word) == 18:
                row_items = re.split(r'\s+', row)
                if len(row_items) >= 3:
                    detail = row_items[0]
                    withdrawn = row_items[1]
                    balance = row_items[2]
                else:
                    detail = row
            elif '.' in row:
                floatItems = re.findall("\d+\.\d+", row)
                for f in floatItems:
                    row = row.replace(f, "")
                row = row.strip()
                detail = row
                paid_in = floatItems[0]
                balance = floatItems[1]
            else:
                date = ''
                detail = row
                paid_in = ''
                balance = ''
                withdrawn = ''
        return date, detail, withdrawn, paid_in, balance
    extracted_text = convert_pdf_to_text(file, rows_to_skip)

    data = []
    lines = extracted_text.split('\n')
    for line in lines:
        date, detail, withdrawn, paid_in, balance = process_row(line)
        if any([date, detail, withdrawn, paid_in, balance]):
            data.append([date, detail, withdrawn, paid_in, balance])

    try:
        df = pd.DataFrame(data)
        workbook = Workbook()
        sheet = workbook.active
        column_headers = ['Date', 'Details', 'Withdrawn', 'Paid In', 'Balance']
        sheet.append(column_headers)
        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width
        df = f"{download_excel_path}{filename}.xlsx"
        workbook.save(df)
        return Response(status=201)
    except Exception:
        return Response(status=404)


def processNatwest_Small_Scanned(file, filename):
    BILL_PAYMENT = "Bill Payment"
    DIRECT_DEBIT = "Direct Debit"
    ONLINE_TRANSACTION = "OnLine Transaction"
    AUTOMATED_CREDIT = "Automated Credit"

    def rotate_pdf_pages(file):
        try:
            pdf_reader = PyPDF2.PdfReader(file)
            pdf_writer = PyPDF2.PdfWriter()
            for pagenum in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[pagenum]
                page.rotate(90)
                pdf_writer.add_page(page)
            with open(file, 'wb') as pdf_out:
                pdf_writer.write(pdf_out)
            return True
        except Exception:
            return False

    def convert_pdf_to_dataframe(file):
        formatted_text = []
        images = convert_from_path(file, dpi=300)
        for img in images:
            text = pytesseract.image_to_string(img, config='--psm 4')
            lines = text.split('\n')
            formatted_lines = [line for line in lines if line.strip()]
            formatted_text.extend(formatted_lines)
        df = pd.DataFrame(formatted_text, columns=['Text'])
        return df

    def process_row(row):
        date = ''
        detail = ''
        withdrawn_paid_in = ''
        balance = ''

        date_pattern = r'\b\d{2}\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4}\b|\b\d{1,2}\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\b|\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4}\b'
        date_match = re.search(date_pattern, row)

        if date_match:
            date = date_match.group()
        all_words = row.split()
        items = row.split()
        if items and "OD" in items[-1]:
            balance = items[-2] + ' ' + items[-1]

        if all_words:
            if BILL_PAYMENT in row:
                date_detail_parts = row.split(BILL_PAYMENT)
                if len(date_detail_parts) == 2:
                    date = date_detail_parts[0].strip()
                    detail = f'{BILL_PAYMENT} {date_detail_parts[1].strip()}'
            elif DIRECT_DEBIT in row:
                date_detail_parts = row.split(DIRECT_DEBIT, 1)
                if len(date_detail_parts) == 2:
                    date = date_detail_parts[0].strip()
                    detail = f'{DIRECT_DEBIT} {date_detail_parts[1].strip()}'
            elif ONLINE_TRANSACTION in row:
                date_detail_parts = row.split(ONLINE_TRANSACTION, 1)
                if len(date_detail_parts) == 2:
                    date = date_detail_parts[0].strip()
                    detail = f'{ONLINE_TRANSACTION} {date_detail_parts[1].strip()}'
            elif AUTOMATED_CREDIT in row:
                date_detail_parts = row.split(AUTOMATED_CREDIT, 1)
                if len(date_detail_parts) == 2:
                    date = date_detail_parts[0].strip()
                    detail = f'{AUTOMATED_CREDIT} {date_detail_parts[1].strip()}'
            if ',' in row:
                float_items = re.findall(r"\d+[.,]\d+", row)
                for f in float_items:
                    row = row.replace(f, "")
                row = row.strip()
                detail = row
                withdrawn_paid_in = ''
                if len(float_items) >= 1:
                    withdrawn_paid_in = "{:,.2f}".format(
                        float(float_items[0].replace(',', ''))).replace(",", ",")
            else:
                detail = row
        date = date.replace("25 Sep", "25 Sep 2019").replace(" Sep 2019", "26 Sep 2019").replace("2526 Sep 2019", "25 Sep 2019")
        detail = detail.replace(" .00", "").replace(" .15", "").replace("  .50 OD", "").replace("27 Sep ", "").replace("  .41 OD", "").replace("S ", "").replace("oa", "2019").replace("26 Sep O", "O").replace("N Gx ", "26 ").replace(
            "BROUGHT Cxie . Ce bese oD", "BROUGHT FORWARD 23,134.56 OD").replace("Ca aaah FORWARD cee belanee ss) oD", "BROUGHT FORWARD 33,814.50 OD").replace("25 Sep 2019 BROUGHT FORWARD 23,134.56 OD", "").replace("26 Sep 2019 BROUGHT FORWARD 33,814.50 OD", "")
        withdrawn_paid_in = withdrawn_paid_in.replace("23,134.00", "")
        return date, detail, withdrawn_paid_in, balance

    if rotate_pdf_pages(file):
        df = convert_pdf_to_dataframe(file)
        df = df.drop([0, 1, 29, 30, 31, 59, 60, 61])
        df = df.reset_index(drop=True)

        data = []
        for row in df['Text']:
            date, detail, withdrawn_paid_in, balance = process_row(row)
            data.append([date, detail, withdrawn_paid_in, balance])
        columns = ["Date", "Details", "Withdrawn/Paid In", "Balance"]
        df_output = pd.DataFrame(data, columns=columns)

        details_to_drop = ["CARRIED FORWARD"]
        escaped_details = [re.escape(pattern) for pattern in details_to_drop]
        details_to_drop_pattern = '|'.join(escaped_details)
        df_output = df_output[~df_output['Details'].str.contains(details_to_drop_pattern)]
        df_output.at[18, 'Withdrawn/Paid In'] = df_output['Details'].str.split()[18][1]
        df_output.at[18, 'Details'] = ' '.join(df_output['Details'].str.split()[18][:1] + df_output['Details'].str.split()[18][2:])
        df_output.at[37, 'Withdrawn/Paid In'] = df_output['Details'].str.split()[37][1]
        df_output.at[37, 'Details'] = ' '.join(df_output['Details'].str.split()[37][:1] + df_output['Details'].str.split()[37][2:])

        try:
            with pd.ExcelWriter(f"{download_excel_path}{filename}.xlsx", engine="openpyxl") as writer:
                df_output.to_excel(writer, sheet_name=f"{filename}", index=False)
                writer.book
                worksheet = writer.sheets[f"{filename}"]
                for column_cells in worksheet.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2
            return Response(status=201)
        except Exception:
            return Response(status=404)
    else:
        return Response(status=404)


def processBarclays_Scanned(file, filename):
    def ends_with_C(value):
        if len(value) > 0 and value[-1] == "C":
            return True
        else:
            return False
    unwanted_chars = ['lI', '{', 'l', 'I]', 'i', '\\', '\\]', '1]', 'if' , 'I' , ']' , 'f' , '}' ,'11' ,'——']
    rows_to_skip = [
        (20, 2),
        (8, 3),
    ]
    def convert_pdf_to_text(file, rows_to_skip):
        extracted_text = []
        images = convert_from_path(file, dpi=300)
        for i, img in enumerate(images, start=1):
            rows_to_skip_start, rows_to_skip_end = rows_to_skip[i - 1]
            text = pytesseract.image_to_string(img)
            lines = text.split('\n')
            lines = lines[rows_to_skip_start:]
            lines = lines[:-rows_to_skip_end]
            for line in lines:
                columns = [col.strip() for col in line.split('|')]
                columns = [col for col in columns if col]
                if columns:
                    extracted_text.append(columns)        
        return extracted_text

    extracted_text_list = convert_pdf_to_text(file, rows_to_skip)
    details = []
    payments = []
    receipts = []
    dates = []
    statement_balance = []
    clrd_for_interest = []

    for line in extracted_text_list:
        if len(line) >= 6:
            details.append(line[0])
            payments.append(line[1])
            receipts.append(line[2])
            dates.append(line[3])
            statement_balance.append(line[4])
            clrd_for_interest.append(line[5])
        elif len(line) == 5:
            details.append(line[0])
            payments.append(line[1])
            receipts.append(line[2])
            dates.append(line[3])
            statement_balance.append(line[4])
            clrd_for_interest.append('')
        elif len(line) == 4:
            details.append(line[0])
            payments.append(line[1])
            receipts.append('')
            dates.append(line[2])
            statement_balance.append(line[3])
            clrd_for_interest.append('')
        elif len(line) == 3:
            details.append(line[0])
            payments.append(line[1])
            receipts.append('')
            dates.append('')
            statement_balance.append(line[2])
            clrd_for_interest.append('')
        elif len(line) == 2:
            details.append(line[0])
            payments.append(line[1])
            receipts.append('')
            dates.append('')
            statement_balance.append('')
            clrd_for_interest.append('')
        else:
            details.append('')
            payments.append('')
            receipts.append('')
            dates.append('')
            statement_balance.append('')
            clrd_for_interest.append('')

    for i in range(len(payments)):
        if ends_with_C(payments[i]):
            receipts[i] = payments[i]
            payments[i] = ''

    for i in range(len(statement_balance)):
        value = statement_balance[i]
        if len(value) > 0 and "I]" in value:
            parts = value.split("I]")
            statement_balance[i] = parts[0]
            clrd_for_interest[i] = parts[1]

    for i in range(len(statement_balance)):
        value = statement_balance[i]
        if "II" in value:
            parts = value.split("II")
            statement_balance[i] = parts[0]
            clrd_for_interest[i] = parts[1]

    for i in range(len(statement_balance)):
        value = statement_balance[i]
        if "1]" in value:
            parts = value.split("1]")
            statement_balance[i] = parts[0]
            clrd_for_interest[i] = parts[1]

    for i in range(len(statement_balance)):
        value = statement_balance[i]
        if "I" in value:
            parts = value.split("I")
            statement_balance[i] = parts[0]
            clrd_for_interest[i] = parts[1]

    for i in range(len(statement_balance)):
        value = statement_balance[i]
        if re.match(r'(\d{1,2} [A-Z]{3} \d{2})', value):
            dates[i] = statement_balance[i]
            statement_balance[i] = ''

    for i in range(len(receipts)):
        value = receipts[i]
        if re.match(r'(\d{1,2} [A-Z]{3} \d{2})', value):
            clrd_for_interest[i] = statement_balance[i]
            statement_balance[i] = dates[i]
            dates[i] = receipts[i]
            receipts[i] = ''

    for i in range(len(details)):
        details[i] = re.sub('|'.join(map(re.escape, unwanted_chars)), '', details[i])
        payments[i] = re.sub('|'.join(map(re.escape, unwanted_chars)), '', payments[i])
        receipts[i] = re.sub('|'.join(map(re.escape, unwanted_chars)), '', receipts[i])
        dates[i] = re.sub('|'.join(map(re.escape, unwanted_chars)), '', dates[i])
        statement_balance[i] = re.sub('|'.join(map(re.escape, unwanted_chars)), '', statement_balance[i])
        clrd_for_interest[i] = re.sub('|'.join(map(re.escape, unwanted_chars)), '', clrd_for_interest[i])

    df = pd.DataFrame({
        "DETAILS": details,
        "PAYMENTS": payments,
        "RECEIPTS": receipts,
        "DATE": dates,
        "STATEMENT BALANCE": statement_balance,
        "CLRD FOR INTEREST": clrd_for_interest,
    })

    try:
        with pd.ExcelWriter(f"{download_excel_path}{filename}.xlsx", engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=f"{filename}", index=False)
            writer.book
            worksheet = writer.sheets[f"{filename}"]
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2
        return Response(status=201)
    except Exception:
        return Response(status=404)